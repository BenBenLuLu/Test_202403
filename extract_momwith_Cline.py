"""
MoM Raw Extraction Script
=========================
Phase 1 ONLY: Reads all .docx files from MOM_FOLDER and exports
mon_rawtable.xlsx with three sheets:

  Raw Table       -- every paragraph / table row with Section tags
  Headers         -- one row per file with all pre-parsed header fields
  Copilot_Prompt  -- ready-to-paste prompt for Microsoft Copilot

Output columns (Raw Table):
  File_Index | Source_File | Seq | Type | Style | Section | Content | Has_Image

Output columns (Headers):
  File_Index | 檔名 | Subject | No | Date | Place | Recorded_by | Approved_by
  | Issue_Type | Priority | Counter_Previous | Counter_Current

No AI / Ollama required.

Usage (PowerShell on Windows):
  & "C:\\Users\\ben.lu\\OneDrive - shl-group.com\\Documents\\Privacy\\python2\\Ben\\AI agent\\.venv\\Scripts\\python.exe" ^
    "C:\\Users\\ben.lu\\OneDrive - shl-group.com\\Documents\\Privacy\\python2\\Ben\\SHL\\Tool Assessment\\MoM extraction\\extract_momwith_Cline.py"
"""

# =============================================================================
# Last commit : 4576eb0  (2026-06-26)
# Branch      : cursor/phase1-rawonly-66b9
# Note        : Phase 1 only -- no Ollama / AI dependency
# =============================================================================

import sys
import io
import subprocess
import re
from pathlib import Path

# Make stdout safe on cp950 / cp936 terminals
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(errors="replace")
else:
    sys.stdout = io.TextIOWrapper(
        sys.stdout.buffer,
        encoding=sys.stdout.encoding or "utf-8",
        errors="replace",
    )

# ── Dependency auto-install ───────────────────────────────────────────────────

def _ensure_package(import_name: str, pip_name: str | None = None) -> None:
    try:
        __import__(import_name)
    except ImportError:
        pkg = pip_name or import_name
        print(f"[setup] '{pkg}' not found -- installing...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg])
        print(f"[setup] '{pkg}' installed.")


_ensure_package("docx",    "python-docx")
_ensure_package("pandas")
_ensure_package("openpyxl")

import docx           # noqa: E402
import pandas as pd   # noqa: E402

# ── Configuration ─────────────────────────────────────────────────────────────

MOM_FOLDER  = Path(r"c:\Users\ben.lu\OneDrive - shl-group.com\Documents\Privacy\python2\Ben\SHL\Tool Assessment\MoM extraction\MoM")
OUTPUT_XLSX = MOM_FOLDER / "mon_rawtable.xlsx"

# ── Text normalisation ────────────────────────────────────────────────────────

_UNICODE_REPLACEMENTS = [
    # Arrows
    ("\u2192", "->"),   # RIGHT ARROW
    ("\u21D2", "->"),   # RIGHTWARDS DOUBLE ARROW
    ("\u27A1", "->"),   # BLACK RIGHTWARDS ARROW
    ("\u2794", "->"),   # HEAVY WIDE-HEADED RIGHTWARDS ARROW
    ("\u279C", "->"),   # HEAVY ROUND-TIPPED RIGHTWARDS ARROW
    ("\u2190", "<-"),   # LEFT ARROW
    ("\u2194", "<->"),  # LEFT RIGHT ARROW
    # Dashes / bullets
    ("\u2013", "-"),    # en dash
    ("\u2014", "--"),   # em dash
    ("\u2022", "*"),    # bullet
    ("\u25CF", "*"),    # black circle
    # Checkboxes
    ("\u2611", "[x]"),  # checked ballot box
    ("\u2612", "[x]"),  # ballot box with X
    ("\u2610", "[ ]"),  # empty ballot box
    ("\u25A0", "[x]"),  # black square  (filled checkbox)
    ("\u25A1", "[ ]"),  # white square  (empty checkbox)
    ("\u25AA", "[x]"),  # black small square
    # Misc
    ("\u2026", "..."),  # ellipsis
    ("\u00D7", "x"),    # multiplication sign
    ("\u00B0", "deg"),  # degree sign
]

def _normalize(text: str) -> str:
    """Replace known Unicode symbols with ASCII equivalents."""
    for char, replacement in _UNICODE_REPLACEMENTS:
        if char in text:
            text = text.replace(char, replacement)
    return text

# ── Section & field detection helpers ────────────────────────────────────────

# Maps section heading keywords -> canonical Section name
_SECTION_KEYWORDS = {
    "agenda":       "Agenda",
    "root cause":   "Root_Cause",
    "solution":     "Solutions",
    "solutions":    "Solutions",
    "action item":  "Action_Items",
    "action items": "Action_Items",
}

_CHECKBOX_ISSUE_KW    = ["fot issue", "ts issue", "production issue"]
_CHECKBOX_PRIORITY_KW = ["super urgent", "urgent", "general"]

# Maps "label:" prefixes in header table cells to column names
_HEADER_LABEL_MAP = {
    "subject":          "Subject",
    "no":               "No",
    "no.":              "No",
    "date":             "Date",
    "place":            "Place",
    "location":         "Place",
    "recorded by":      "Recorded_by",
    "recorder":         "Recorded_by",
    "approved by":      "Approved_by",
    "approver":         "Approved_by",
    "previous count":   "Counter_Previous",
    "counter previous": "Counter_Previous",
    "current count":    "Counter_Current",
    "counter current":  "Counter_Current",
}


def _section_from_text(text: str) -> str | None:
    """Return canonical section name if *text* is a section heading, else None."""
    norm = re.sub(r"^\s*[\d\w]+[\.\)]\s*", "", text.strip()).lower()
    for kw, section in _SECTION_KEYWORDS.items():
        if norm == kw or norm.startswith(kw + ":") or norm.startswith(kw + " "):
            return section
    return None


def _is_checkbox_paragraph(text: str) -> bool:
    """True when the paragraph is the issue-type / priority checkbox line."""
    low = text.lower()
    return ("[x]" in low or "[ ]" in low) and any(
        kw in low for kw in _CHECKBOX_ISSUE_KW + _CHECKBOX_PRIORITY_KW
    )


def _classify_table_section(
    all_cells_text: str,
    table_count: int,
    current_section: str,
) -> str:
    """Decide which Section to assign to a whole table block."""
    low = all_cells_text.lower()
    if table_count == 1:
        return "Header"
    if "previous count" in low or "current count" in low:
        return "Counter"
    if table_count == 2 and current_section == "Unknown":
        return "Header"          # second header table (e.g. urgency / counter)
    if current_section in ("Solutions", "Action_Items", "Root_Cause"):
        return current_section   # table embedded inside a known section
    return "Unknown"


def _parse_checkbox_fields(content: str) -> tuple[str, str]:
    """Extract Issue_Type and Priority from a normalised checkbox paragraph.
    Returns (issue_type, priority) — empty string if not found.
    """
    issue_type = ""
    priority   = ""
    for m in re.finditer(r"\[x\]\s*([^\[\]]+)", content, re.IGNORECASE):
        item = m.group(1).strip().lower()
        for kw in _CHECKBOX_ISSUE_KW:
            if item.startswith(kw) and not issue_type:
                issue_type = item[:len(kw)].strip()
        for kw in _CHECKBOX_PRIORITY_KW:
            if item.startswith(kw) and not priority:
                priority = item[:len(kw)].strip()
    return issue_type, priority


def _split_label_value(cell: str) -> tuple[str, str]:
    """Split 'Label: Value' at the first colon. Returns ('', cell) if none."""
    for sep in (": ", ":"):
        idx = cell.find(sep)
        if 0 < idx < len(cell) - len(sep):
            return cell[:idx].strip(), cell[idx + len(sep):].strip()
    return "", cell.strip()

# ── Phase 1: Extract .docx → flat rows ───────────────────────────────────────

def extract_rows_from_docx(filepath: Path, file_index: int) -> list[dict]:
    """
    Walk the XML body in document order, tagging each row with a Section.

    Returns one dict per paragraph or table row:
        file_index  : int   -- 1-based index of this file in the batch
        source_file : str   -- filename only
        seq         : int   -- 1-based sequence within this file
        type        : str   -- "paragraph" | "table_row"
        style       : str   -- Word paragraph style; "table" for table rows
        section     : str   -- Header | Counter | Checkbox | Agenda |
                               Root_Cause | Solutions | Action_Items | Unknown
        content     : str   -- normalised extracted text
        has_image   : bool  -- True if document has any inline image
    """
    from docx.oxml.ns import qn  # noqa: PLC0415

    doc       = docx.Document(str(filepath))
    has_image = len(doc.inline_shapes) > 0
    rows: list[dict] = []
    seq             = 1
    table_count     = 0
    current_section = "Unknown"

    for child in doc.element.body.iterchildren():
        tag = child.tag.split("}")[-1] if "}" in child.tag else child.tag

        # ── Paragraph ──────────────────────────────────────────────────────
        if tag == "p":
            text = _normalize(
                "".join(run.text for run in child.iter(qn("w:t"))).strip()
            )
            if not text:
                continue
            pStyle = child.find(".//" + qn("w:pStyle"))
            style  = pStyle.get(qn("w:val")) if pStyle is not None else "Normal"

            if _is_checkbox_paragraph(text):
                section = "Checkbox"
            else:
                new_sec = _section_from_text(text)
                if new_sec:
                    current_section = new_sec
                section = current_section

            rows.append({
                "file_index":  file_index,
                "source_file": filepath.name,
                "seq":         seq,
                "type":        "paragraph",
                "style":       style,
                "section":     section,
                "content":     text,
                "has_image":   has_image,
            })
            seq += 1

        # ── Table ───────────────────────────────────────────────────────────
        elif tag == "tbl":
            table_count += 1
            # Collect all cells first so we can classify the whole table
            tbl_data: list[list[str]] = []
            for row_el in child.iter(qn("w:tr")):
                cells = [
                    _normalize(
                        "".join(t.text for t in c.iter(qn("w:t"))).strip()
                    )
                    for c in row_el.iter(qn("w:tc"))
                ]
                tbl_data.append(cells)

            all_cells_text = " ".join(" ".join(r) for r in tbl_data)
            tbl_section    = _classify_table_section(
                all_cells_text, table_count, current_section
            )
            # Update current_section only if the table lives inside a body section
            if tbl_section not in ("Header", "Counter", "Unknown"):
                current_section = tbl_section

            for cells in tbl_data:
                line = " | ".join(cells)
                if line.strip(" |"):
                    rows.append({
                        "file_index":  file_index,
                        "source_file": filepath.name,
                        "seq":         seq,
                        "type":        "table_row",
                        "style":       "table",
                        "section":     tbl_section,
                        "content":     line,
                        "has_image":   has_image,
                    })
                    seq += 1

    return rows


def build_raw_dataframe(mom_folder: Path) -> pd.DataFrame:
    """Iterate all .docx files and return combined raw DataFrame."""
    if not mom_folder.exists():
        raise FileNotFoundError(
            f"MoM folder not found:\n  {mom_folder}"
        )
    files = sorted(mom_folder.glob("*.docx"))
    if not files:
        raise ValueError(f"No .docx files found in:\n  {mom_folder}")

    all_rows: list[dict] = []
    for idx, fp in enumerate(files, start=1):
        print(f"  [{idx:>3}] Extracting: {fp.name}")
        try:
            rows = extract_rows_from_docx(fp, idx)
            all_rows.extend(rows)
            sections = {r["section"] for r in rows}
            print(f"         -> {len(rows)} rows | sections: {sorted(sections)}")
        except Exception as exc:  # noqa: BLE001
            print(f"         [WARNING] {exc}")

    df = pd.DataFrame(all_rows).rename(columns={
        "file_index":  "File_Index",
        "source_file": "Source_File",
        "seq":         "Seq",
        "type":        "Type",
        "style":       "Style",
        "section":     "Section",
        "content":     "Content",
        "has_image":   "Has_Image",
    })
    return df[[
        "File_Index", "Source_File", "Seq", "Type",
        "Style", "Section", "Content", "Has_Image",
    ]]

# ── Build Headers sheet ───────────────────────────────────────────────────────

_HEADERS_COLS = [
    "File_Index", "檔名", "Subject", "No", "Date", "Place",
    "Recorded_by", "Approved_by", "Issue_Type", "Priority",
    "Counter_Previous", "Counter_Current",
]


def build_headers_dataframe(raw_df: pd.DataFrame) -> pd.DataFrame:
    """
    One row per file.  Parses header fields from Section=Header/Counter rows
    and checkbox fields from Section=Checkbox rows.
    """
    records: list[dict] = []

    for file_idx, grp in raw_df.groupby("File_Index", sort=True):
        rec = {c: "" for c in _HEADERS_COLS}
        rec["File_Index"] = file_idx
        rec["檔名"]        = grp["Source_File"].iloc[0]

        # Parse Label: Value from Header + Counter rows
        header_rows = grp[grp["Section"].isin(["Header", "Counter"])]
        for _, row in header_rows.iterrows():
            for cell in row["Content"].split(" | "):
                label, value = _split_label_value(cell.strip())
                if not label or not value:
                    continue
                norm = label.lower()
                for pattern, field in _HEADER_LABEL_MAP.items():
                    if norm == pattern or norm.startswith(pattern):
                        if not rec[field]:
                            rec[field] = value
                        break

        # Parse Issue_Type + Priority from Checkbox rows
        checkbox_rows = grp[grp["Section"] == "Checkbox"]
        if not checkbox_rows.empty:
            combined = " ".join(checkbox_rows["Content"].tolist())
            issue, priority = _parse_checkbox_fields(combined)
            rec["Issue_Type"] = issue
            rec["Priority"]   = priority

        records.append(rec)

    return pd.DataFrame(records, columns=_HEADERS_COLS)

# ── Copilot prompt template ───────────────────────────────────────────────────

COPILOT_PROMPT = """\
=== Microsoft Copilot Prompt Template ===
Use this prompt in Excel Copilot or Microsoft 365 Copilot Chat.

------------------------------------------------------------
PROMPT (paste into Copilot):
------------------------------------------------------------
You are analysing mold-tooling Minutes of Meeting (MoM) data
from the two sheets in this workbook:

  "Headers"   -- one row per MoM file (Subject, No, Date, etc.)
  "Raw Table" -- every extracted text row with a Section tag

For each File_Index, build a summary table with these columns:
  File_Index, 檔名, Subject, No, Date, Place,
  Recorded_by, Approved_by, Issue_Type, Priority,
  Counter_Previous, Counter_Current,
  Agenda, Root_Cause, Solutions, Action_Items,
  Action (Replacement / Repair+Welding / Other),
  Spare_Part (PartNo * Qty _ PartName format if found),
  Checkpoint (TRUE if Root_Cause + Solutions + Action_Items all non-empty),
  Has_Image

Rules:
1. Copy Subject / No / Date / Place / Recorded_by / Approved_by /
   Issue_Type / Priority / Counter_Previous / Counter_Current
   directly from the "Headers" sheet.
2. For Agenda:      join Content of rows where Section = "Agenda"
3. For Root_Cause:  join Content of rows where Section = "Root_Cause"
4. For Solutions:   join Content of rows where Section = "Solutions"
5. For Action_Items:join Content of rows where Section = "Action_Items"
6. For Action:      classify from Solutions/Action_Items text:
     - "Replacement"    if replacement / replace / new part mentioned
     - "Repair/Welding" if repair / weld / welding mentioned
     - "Other"          otherwise
7. For Spare_Part:  extract part number + quantity + name if mentioned
8. Checkpoint = TRUE only when Root_Cause AND Solutions AND
   Action_Items are all non-empty strings.
9. Has_Image: TRUE if any row for this file has Has_Image = TRUE.
------------------------------------------------------------

=== Column descriptions ===
File_Index      : Sequential file number (1, 2, 3 ...)
Source_File     : .docx filename
Seq             : Row sequence within the file
Type            : paragraph | table_row
Style           : Word paragraph style (Normal, Heading1, etc.)
Section         : Header | Counter | Checkbox | Agenda |
                  Root_Cause | Solutions | Action_Items | Unknown
Content         : Extracted text (Unicode symbols replaced with ASCII)
Has_Image       : TRUE if the .docx contains inline images
"""

# ── Export to Excel ───────────────────────────────────────────────────────────

# Section -> background fill colour (Excel hex, no #)
_SECTION_COLOURS = {
    "Header":       "BDD7EE",  # light blue
    "Counter":      "DDEBF7",  # very light blue
    "Checkbox":     "E2EFDA",  # light green
    "Agenda":       "FFF2CC",  # light yellow
    "Root_Cause":   "FCE4D6",  # light orange
    "Solutions":    "E2EFDA",  # light green
    "Action_Items": "EAD1DC",  # light pink
    "Unknown":      "F2F2F2",  # light grey
}


def save_excel(
    raw_df: pd.DataFrame,
    headers_df: pd.DataFrame,
    output_path: Path,
) -> None:
    """Write three sheets: Raw Table, Headers, Copilot_Prompt."""
    from openpyxl.styles import Font, PatternFill, Alignment  # noqa: PLC0415
    from openpyxl.utils import get_column_letter               # noqa: PLC0415

    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:

        # ── Sheet 1: Raw Table ────────────────────────────────────────────
        raw_df.to_excel(writer, index=False, sheet_name="Raw Table")
        ws_raw = writer.sheets["Raw Table"]
        ws_raw.freeze_panes = "A2"

        _style_header_row(ws_raw)
        _auto_width(ws_raw, raw_df)

        # Colour each row by Section
        sec_col = list(raw_df.columns).index("Section") + 1
        con_col = list(raw_df.columns).index("Content") + 1
        for ri in range(2, len(raw_df) + 2):
            sec_val = ws_raw.cell(row=ri, column=sec_col).value or "Unknown"
            colour  = _SECTION_COLOURS.get(sec_val, "FFFFFF")
            fill    = PatternFill(start_color=colour, end_color=colour,
                                  fill_type="solid")
            for ci in range(1, len(raw_df.columns) + 1):
                ws_raw.cell(row=ri, column=ci).fill = fill
            ws_raw.cell(row=ri, column=con_col).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

        # ── Sheet 2: Headers ──────────────────────────────────────────────
        headers_df.to_excel(writer, index=False, sheet_name="Headers")
        ws_hdr = writer.sheets["Headers"]
        ws_hdr.freeze_panes = "A2"
        _style_header_row(ws_hdr)
        _auto_width(ws_hdr, headers_df)

        # ── Sheet 3: Copilot_Prompt ───────────────────────────────────────
        from openpyxl import Workbook  # noqa: PLC0415, F401
        ws_prompt = writer.book.create_sheet("Copilot_Prompt")
        for i, line in enumerate(COPILOT_PROMPT.splitlines(), start=1):
            ws_prompt.cell(row=i, column=1, value=line)
        ws_prompt.column_dimensions["A"].width = 90
        ws_prompt.freeze_panes = None


def _style_header_row(ws) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment  # noqa: PLC0415
    fill  = PatternFill(start_color="4472C4", end_color="4472C4",
                        fill_type="solid")
    font  = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = align


def _auto_width(ws, df: pd.DataFrame, cap: int = 50) -> None:
    from openpyxl.utils import get_column_letter  # noqa: PLC0415
    for ci, col in enumerate(df.columns, start=1):
        max_len = max(
            len(str(col)),
            df[col].astype(str).map(len).max() if len(df) else 0,
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, cap)

# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    print("=" * 65)
    print("  MoM Raw Extraction  (Phase 1 only)")
    print("=" * 65)
    print(f"  MoM folder : {MOM_FOLDER}")
    print(f"  Output     : {OUTPUT_XLSX}")
    print()

    print("[Phase 1] Extracting .docx files...")
    raw_df = build_raw_dataframe(MOM_FOLDER)

    n_files = raw_df["File_Index"].nunique()
    n_rows  = len(raw_df)
    print(f"\n  [OK] {n_files} file(s) -> {n_rows} rows total")
    print(f"  Section distribution:\n"
          f"{raw_df['Section'].value_counts().to_string()}")

    print("\n[Building Headers sheet...]")
    headers_df = build_headers_dataframe(raw_df)
    print(headers_df[["File_Index", "檔名", "Subject", "Issue_Type",
                       "Priority"]].to_string(index=False))

    print(f"\n[Saving] {OUTPUT_XLSX} ...")
    save_excel(raw_df, headers_df, OUTPUT_XLSX)

    print(f"\n  [OK] Saved -- {n_rows} raw rows + {len(headers_df)} header rows")
    print(f"  Sheets: Raw Table | Headers | Copilot_Prompt")

    print("\n--- Raw Table preview (first 15 rows) ---")
    preview = raw_df.copy()
    preview["Content"] = (
        preview["Content"].str[:70].str.replace("\n", " ", regex=False)
    )
    print(preview.head(15).to_string(index=False))

    print("\n[Done] Open mon_rawtable.xlsx and use the Copilot_Prompt sheet.")


if __name__ == "__main__":
    main()
